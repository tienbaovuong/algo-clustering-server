import asyncio
import io
from typing import Optional, List
from datetime import datetime

import openpyxl
from fastapi import UploadFile
from beanie import PydanticObjectId
from beanie.operators import RegEx, GTE, Eq, In

from app.helpers.exceptions import NotFoundException, ThesisWrongFormatException
from app.dto.thesis_data_dto import ShortThesisData, FullThesisData
from app.models.thesis_data import ThesisData
from app.worker.tasks.nlp_task import schedule_preprocess


def parse_xlsx_thesis(xlsx_file):
    wb = openpyxl.load_workbook(xlsx_file, read_only=True)
    thesis = wb[wb.sheetnames[0]]
    input_dict = {
        "student_data": {}
    }
    counter = 0
    column_range = [1,2,8,5]
    for row in range(1, 200):
        if counter >= 7:
            break
        if row >= 200:
            break
        for column in column_range:
            current_cell_value = str(thesis.cell(row, column).value)
            if column == 1:
                if not input_dict.get('student_data').get('student_name') and "Họ và tên sinh viên" in current_cell_value:
                    input_dict['student_data']['student_name'] = thesis.cell(row, column+2).value
                    counter += 1
                elif not input_dict.get('title') and "Tên đề tài" in current_cell_value:
                    input_dict['title'] = thesis.cell(row+1, column).value
                    counter += 1
                    break
                elif not input_dict.get('expected_result') and "Sản phẩm kỳ vọng" in current_cell_value:
                    input_dict['expected_result'] = str(thesis.cell(row+1, column).value).replace("-","").replace("\n", "").strip()
                    counter += 1
                    break
                elif not input_dict.get('problem_solve') and "Vấn đề thực tiễn đồ án giải quyết" in current_cell_value:
                    input_dict['problem_solve'] = str(thesis.cell(row+1, column).value).replace("-","").replace("\n", "").strip()
                    counter += 1
                    break
            elif column == 2:
                if not input_dict.get('category') and "Lựa chọn 1" in current_cell_value:
                    input_dict['category'] = thesis.cell(row, column+1).value
                    if thesis.cell(row + 1, column + 1).value:
                        input_dict['category'] += ", " + str(thesis.cell(row+1, column+1).value)
                    column_range.remove(2)
                    counter += 1
                    break
            elif column == 8:
                if not input_dict.get('student_data').get('student_id') and 'MSSV' in current_cell_value:
                    input_dict['student_data']['student_id'] = str(thesis.cell(row, column+1).value)
                    column_range.remove(8)
                    counter += 1
                    break
            elif column == 5:
                if not input_dict.get('semester') and current_cell_value == 'KỲ':
                    input_dict['semester'] = str(thesis.cell(row, column+1).value)
                    column_range.remove(5)
                    counter += 1
                    break    
    wb.close()
    if counter < 7:
        return None
    return input_dict

class ThesisDataService:
    async def list_thesis(
        self,
        title: Optional[str],
        semester: Optional[str],
        created_at: Optional[datetime],
        page: int = 1,
        limit: int = 10,
    ):
        query = []
        skip = limit * (page - 1)
        if title:
            query.append(RegEx(ThesisData.title, title, options="i"))
        if semester:
            query.append(Eq(ThesisData.semester, semester))
        if created_at:
            query.append(GTE(ThesisData.created_at, created_at))
        if limit == -1:
            limit = None
        query_task = ThesisData.find_many(*query)
        total = await query_task.count()
        thesis_data_list = await query_task.skip(skip).limit(limit).sort(-ThesisData.id).project(ShortThesisData).to_list()
        return thesis_data_list, total

    async def get(
        self,
        thesis_id: str
    ):
        thesis_data = await ThesisData.find_one({'_id': PydanticObjectId(thesis_id)}).project(FullThesisData)
        if not thesis_data:
            raise NotFoundException("No thesis")
        return thesis_data

    async def create(
        self,
        file: UploadFile,
    ):
        f = await file.read()
        xlsx_file = io.BytesIO(f)
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(None, parse_xlsx_thesis, xlsx_file)
        print(result)
        if not result:
            raise ThesisWrongFormatException()
        result['updated_at'] = datetime.utcnow()
        model = ThesisData(**result)
        await model.save()
        task_id = schedule_preprocess(model.dict())
        model.nlp_job_id = task_id
        await model.save()
        return str(model.id)

    async def delete(
        self,
        thesis_id: str
    ):
        thesis = await ThesisData.find_one({'_id': PydanticObjectId(thesis_id)})
        if not thesis:
            raise NotFoundException("No thesis")
        await thesis.delete()

    async def get_list_by_ids(
        self,
        list_ids: List[str],
    ):
        new_list = []
        for id in list_ids:
            new_list.append(PydanticObjectId(id))
        thesis_list = await ThesisData.find_many(
            In(ThesisData.id, new_list)
        ).to_list()
        return thesis_list
    
    async def update_nlp(
        self,
        thesis_id: str,
        title_vector: List,
        category_vector: List,
        expected_result_vector: List,
        problem_solve_vector: List,
    ):
        update_data = {
            "title_vector": title_vector,
            "category_vector": category_vector,
            "expected_result_vector": expected_result_vector,
            "problem_solve_vector": problem_solve_vector,
            "updated_at": datetime.utcnow(),
            "need_nlp_extract": False
        }
        await ThesisData.find_one({'_id': PydanticObjectId(thesis_id)}).update({"$set": update_data})

    async def suggest_cluster_name(
        self,
        list_ids: List[str],
    ):
        thesis_list = await self.get_list_by_ids(list_ids)
        category_dict = {}
        for thesis in thesis_list:
            list_category = [word.strip().capitalize() for word in thesis.category.split(',')]
            for cate in list_category:
                if cate in category_dict.keys():
                    category_dict[cate] += 1
                else:
                    category_dict[cate] = 1

        order = sorted(category_dict.items(), key=lambda x:x[1], reverse=True)
        if len(order) > 1:
            return [order[0][0], order[1][0]]
        elif len(order) == 1:
            return [order[0][0]]
        return ["Default cluster name"]